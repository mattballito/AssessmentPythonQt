import openpyxl


# Load the raw_data workbook
wb1 = openpyxl.load_workbook('raw_data.xlsx')

# Load the regression_data workbook
wb2 = openpyxl.load_workbook('regression_data.xlsx')

ws1, ws2 = wb1.active, wb2.active

# Load output spreadsheet
# Create a new workbook & add a worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write some data to the cells
ws['A1'] = 'analytical_result_id'
ws['B1'] = 'response(x)'
ws['C1'] = 'Maximizing m(slope)'
ws['D1'] = 'Maximizing y_intercept(b)'
ws['E1'] = 'Maximum A'
ws['F1'] = 'Maximizing Analyte ID'


"""given a certain response (x) from raw_data, calculates the greatest A value using data from regression_data (a = mx+b),
   then returns the analyte that causes the maximum A
"""
def getMaxAnalyte(response):
    maxVal = 0
    maxAnalyte = None
    maxSlope = None
    maxIntercept = None
    next_row = ws.max_row # write on the same row that was started earlier

    for row_regression in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        values = tuple([cell.value for cell in row_regression])
        analyte_id, slope, intercept = values[0], values[1], values[2]
        a = (slope*response) + intercept        
            
        if (a == max(a, maxVal)):
            maxAnalyte = analyte_id
            maxVal = a
            maxIntercept = intercept
            maxSlope = slope



    ws['C{}'.format(next_row)] = maxSlope
    ws['D{}'.format(next_row)] = maxIntercept
    ws['E{}'.format(next_row)] = maxVal
    ws['F{}'.format(next_row)] = maxAnalyte
    return maxAnalyte



# Iterate through the rows of the worksheet
for row_raw in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
   
   # Get the values for each cell in the row
   values = tuple([cell.value for cell in row_raw])
   result_id, response = values[0], values[1]

   # Write to the next available row in the output sheet
   next_row = ws.max_row + 1
   ws['A{}'.format(next_row)] = result_id
   ws['B{}'.format(next_row)] = response
   

   
   # from regression_data, find analyte w/ greatest a (a = mx+b)
   result_largest = (result_id,getMaxAnalyte(response))
   print(result_largest)

wb.save('output_openpyxl.xlsx')



        



